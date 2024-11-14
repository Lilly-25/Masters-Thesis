import pandas as pd
import openpyxl
import re
import os
import csv
import numpy as np
import json

def create_tabular_data(file_path, purpose):
    try:
        if os.path.basename(file_path) == '.gitkeep':
            print("Skipping .gitkeep file.")
            return None
        df = pd.read_excel(file_path, sheet_name='input_data', header=None)
        df = df.dropna(how='all').dropna(axis=1, how='all')
        file_name = os.path.splitext(os.path.basename(file_path))[0]
        params_dict = {}
        for _, row in df.iterrows():
            param, value = row[0], row[1]
            if pd.notna(param) and pd.notna(value):
                try:
                    params_dict[param] = float(value)
                except ValueError:
                    params_dict[param] = value
                    
        # Save parameters to a JSON file
        params_file = './data/EMTabular.json'


        # Load parameters from the JSON file
        with open(params_file, 'r') as f:
            params = json.load(f)

        params_rotor_v = params["params_rotor_v"]
        params_rotor_delta = params["params_rotor_delta"]
        params_stator = params["params_stator"]
        params_general = params["params_general"]
 
        df_features=pd.DataFrame()
        
        #V represent max double v magnets and b max delta magnets
        v=2
        b=1
        while v >= 1:
            for key in params_rotor_v:
                df_features.loc[file_name, f"{key}{v}"] = params_dict.get(f"{key}{v}", 0)
            v -= 1
        while b >= 1:
            for key in params_rotor_delta:
                df_features.loc[file_name, f"{key}{b}"] = params_dict.get(f"{key}{b}", 0)
            b -= 1
        for key in params_stator:
            df_features.loc[file_name, f"{key}"] = params_dict.get(f"{key}", 0)
        for key in params_general:
            df_features.loc[file_name, f"{key}"] = params_dict.get(f"{key}", 0)
        
        # Filter all 'deg_phi' columns and create new columns with their corresp radian values
        deg_columns = df_features.filter(regex='^deg_phi').columns

        for col in deg_columns:
            new_col_name = col.replace('deg_', 'rad_')
            df_features[new_col_name] = np.radians(df_features[col])
            
        df_features = df_features.drop(columns=deg_columns)
        #print(df_features.head())
        wb = openpyxl.load_workbook(file_path)
        sheet_mgrenz = wb['Mgrenz']
        mgrenz_values = [cell.value for cell in sheet_mgrenz[1] if cell.value is not None]
        max_mgrenz = max(mgrenz_values)
            
        columns = [f'Column_{i + 1}' for i in range(len(mgrenz_values))]
  
        df_targets = pd.DataFrame(columns=columns)
        df_targets.loc[file_name] = mgrenz_values
        
        max_cols=191
       
        if len(mgrenz_values)!= max_cols:
            print('We have a problem Houston!')
            
        if purpose=='train':
            ##Checking MM sheet to get the correct grid for KPI ETA    
            sheet_mm = wb['MM']

            #Finding out correct indices of the ETA grid
            def check_rows(sheet, start_row, end_row, mgrenz):
                for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
                    first_cell_value = row[0].value  # Get the value of the first cell in the row
                    if first_cell_value == mgrenz:
                        index = row[0].row
                        return index
                

            # Check the first 5 rows
            min_index = check_rows(sheet_mm, 1, 5, -max_mgrenz)

            # Check the last 5 rows
            last_row = sheet_mm.max_row
            max_index = check_rows(sheet_mm, last_row - 4, last_row, max_mgrenz)
            
            #load the eta grid
            sheet_eta = wb['ETA']
            eta_grid_folder = './data/TabularDataETAgrid/'
            if not os.path.exists(eta_grid_folder):
                os.makedirs(eta_grid_folder)
            eta_grid = os.path.join(eta_grid_folder, f"{file_name}.csv")##Instead of saving each grid into a file appeand to a numpy array and save as npy file can access test_size only ased on index
            ##think of an alternative way where we can also have filenames indexed into the numpy array or whatever pythonic object
            with open(eta_grid, mode='w', newline="") as file:
                writer = csv.writer(file)
                mid_eta = False
                for row in sheet_eta.iter_rows(min_row=min_index, max_row=max_index, values_only=True):#Only retrieve the positive ETA grid
                    if not mid_eta:
                        if all(value == 0 for value in row):
                            mid_eta = True
                        else:
                            continue # Ignore the negative ETA grid
                    writer.writerow(row)
        
        return df_features, df_targets            
                
    except Exception as e:
        print(f"Error processing {os.path.basename(file_path)}: {str(e)}")
        return None
        