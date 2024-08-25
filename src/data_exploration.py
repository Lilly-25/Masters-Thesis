import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import patheffects

def read_files_kpi2d(file_path, kpi2):
    ## For 2d KPI we require only 1 row of NN to plot the 2d graph
    wb = openpyxl.load_workbook(file_path)

    sheet_mgrenz = wb[kpi2]
    mgrenz_values = [cell.value for cell in sheet_mgrenz[1] if cell.value is not None]

    sheet_nn = wb['NN']
    nn_values = [cell.value for cell in sheet_nn[1] if cell.value is not None]

    min_length = min(len(mgrenz_values), len(nn_values))
    mgrenz_values = mgrenz_values[:min_length]
    nn_values = nn_values[:min_length]
    
    return nn_values, mgrenz_values


def read_file_kpi3d(file_path, sheet_name):
    # For 3d KPI we need the whole matrix of all 3 dimensions
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    data = [[cell.value if cell.value is not None else np.nan for cell in row] for row in sheet.iter_rows()]
    return np.array(data, dtype=float)

# def plot_kpi2d(nn_values, mgrenz_values):
#     plt.figure(figsize=(10, 6))
#     plt.plot(nn_values, mgrenz_values, marker='o')
#     plt.xlabel('NN Values')
#     plt.ylabel('Mgrenz Values')
#     plt.title(f'NN vs Mgrenz')
#     plt.grid(True)
#     plt.tight_layout()
#     plt.show()
    

def plot_kpi3d(nn, mm, eta):
    fig, ax = plt.subplots(figsize=(16, 10))

    # # Remove the first row and column if headers
    X = nn[0, 1:]
    Y = mm[1:, 0]
    Z = eta[1:, 1:]
    
    # X = nn
    # Y = mm
    # Z = eta[1:, 1:]

    X, Y = np.meshgrid(X, Y)

    mask = np.isfinite(Z)
    X, Y, Z = X[mask], Y[mask], Z[mask]

    contour = ax.tricontourf(X.ravel(), Y.ravel(), Z.ravel(), levels=20, cmap='jet')
    lines = ax.tricontour(X.ravel(), Y.ravel(), Z.ravel(), levels=10, colors='black', linewidths=0.5)
    
    ax.clabel(lines, inline=True, fontsize=8, fmt='%.2f', colors='white',
               inline_spacing=3, use_clabeltext=True)
    
    for text in ax.findobj(match=plt.Text):
        text.set_path_effects([patheffects.withStroke(linewidth=1, foreground='black')])

    ax.set_xlabel('Angular Velocity [rpm]', fontsize=12)
    ax.set_ylabel('Torque [Nm]', fontsize=12)
    ax.set_title('Motor Efficiency', fontsize=14)

    cbar = fig.colorbar(contour)
    cbar.set_label('Efficiency', fontsize=12)

    ax.xaxis.set_major_locator(plt.MaxNLocator(10))
    plt.xticks(rotation=45, ha='right')
    plt.subplots_adjust(bottom=0.15)

    plt.show()
    
def plot_kpi2d(nn_values, mgrenz_values, predicted_mgrenz):
    plt.figure(figsize=(10, 6))
    plt.plot(nn_values, mgrenz_values, marker='o', label='Target', color='blue')
    plt.plot(nn_values, predicted_mgrenz, marker='o', label='Predictions', color='red')
    plt.xlabel('NN Values')
    plt.ylabel('Mgrenz Values')
    plt.title(f'NN vs Mgrenz')
    plt.grid(True)
    plt.tight_layout()
    plt.legend()
    plt.show()