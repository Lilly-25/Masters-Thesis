import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from tqdm import tqdm
import os
import pandas as pd
from matplotlib.colors import Normalize
import matplotlib.ticker as ticker
from matplotlib.cm import ScalarMappable

def read_file_1d(file_path, sheet_name):
    # We only need the 1st row from NN and 1st column from MM for plotting
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    if sheet_name == 'MM':
        data = [cell.value for cell in sheet['A'] if cell.value is not None]
    else:
        data = [cell.value for cell in sheet[1] if cell.value is not None]
    return data

def read_file_eta(file_path, sheet_name):
    # For 3d KPI we need the whole matrix of all 3 dimensions
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    data = [[cell.value if cell.value is not None else np.nan for cell in row] for row in sheet.iter_rows()]
    return np.array(data, dtype=float)

def read_file_2d(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    
    y2_pos = []
    mid_eta = False
    
    for row in sheet.iter_rows():
        if not mid_eta:
            if all(cell.value == 0 for cell in row):
                mid_eta = True
            else:
                continue  # Ignore the negative ETA grid
        
        row_data = [cell.value if cell.value is not None else np.nan for cell in row]
        y2_pos.append(row_data)
    
    # y2 = []
    # y2_neg = y2_pos[-1:0:-1] # Mirroring negative grid of ETA to be same as the predicted positive grid excl mid row
    # y2 = y2_neg + y2_pos
    # while len(y2) < sheet.max_row:
    #     y2.append([np.nan] * len(y2[0]))
    return np.array(y2_pos, dtype=float)
        
        
def plot_kpi2d(nn_values, mgrenz_values):
    plt.figure(figsize=(8, 6))
    plt.plot(nn_values, mgrenz_values, color='blue')
    # plt.plot(nn_values, mgrenz_values, label='Target', color='blue')
    # plt.plot(nn_values, predicted_mgrenz, label='Predictions', color='red')
    plt.xlabel('Torque (Nm)', fontsize=12)
    plt.ylabel('Angular Velocity (rpm)', fontsize=12)
    # plt.title(f'Torque Curve')
    plt.xlim(left=0)  
    plt.ylim(bottom=0)  
    ax=plt.gca()
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(lambda x, pos: f"{int(x)}" if x != 0 else ""))
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    # plt.legend()
    plt.savefig('./Manuscript/ReportImages/TorqueCurve.png', bbox_inches='tight')


def plot_kpi3d(nn, mm, eta):
    
    fig, ax = plt.subplots(figsize=(10, 6))

    Z_global_min = 0.00
    Z_global_max = 100.00
    norm = Normalize(vmin=Z_global_min, vmax=Z_global_max)

    X, Y = np.meshgrid(nn, mm)
    Z = eta

    im = ax.pcolormesh(X, Y, Z, cmap='jet', norm=norm, shading='auto')

    ax.set_xlabel('Angular Velocity (rpm)', fontsize=12)
    ax.set_ylabel('Torque (Nm)', fontsize=12)
    # ax.set_title('Efficiency Grid', fontsize=14)
    
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(lambda x, pos: f"{int(x)}" if x != 0 else ""))
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    cbar = fig.colorbar(im, ax=ax)
    cbar.set_label('Efficiency (%)', fontsize=12)
    cbar.ax.tick_params(labelsize=12) 

    ax.xaxis.set_major_locator(plt.MaxNLocator(10))
    # plt.xticks(rotation=45, ha='right')
    plt.subplots_adjust(bottom=0.15)
    plt.savefig('./Manuscript/ReportImages/EfficiencyGrid.png', bbox_inches='tight')
    
    
def remove_faulty_files(directory):
    # Loop over all files to check if there are any faulty files and remove them
    for filename in tqdm(os.listdir(directory)):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(directory, filename)
            try:
                xls = pd.ExcelFile(file_path)
                sheet_names = xls.sheet_names
                if "ETA" not in sheet_names or "MM" not in sheet_names or "input_data" not in sheet_names or "NN" not in sheet_names or "Mgrenz" not in sheet_names:
                    print(f"{filename} is missing required sheets.")
                    os.remove(file_path)
                    print(f"{filename} removed.")
            except Exception as e:
                print(f"Error reading {filename}: {e}")
                

def artifact_deletion():
    wandb_folder = os.path.expanduser('~/.local/share/wandb')
    if os.path.exists(wandb_folder):
        for root, dirs, files in os.walk(wandb_folder, topdown=False):
            for name in files:
                os.remove(os.path.join(root, name))
            for name in dirs:
                os.rmdir(os.path.join(root, name))
        os.rmdir(wandb_folder)
        print(f"Deleted the folder: {wandb_folder}")
    wandb_folder = os.path.expanduser('~/.cache/wandb')
    if os.path.exists(wandb_folder):
        for root, dirs, files in os.walk(wandb_folder, topdown=False):
            for name in files:
                os.remove(os.path.join(root, name))
            for name in dirs:
                os.rmdir(os.path.join(root, name))
        os.rmdir(wandb_folder)
        print(f"Deleted the folder: {wandb_folder}")
    else:
        print(f"The folder {wandb_folder} does not exist.")
        
        
def cumulative_counts(arr):
    if len(arr) == 0:
        return []
    
    result = []
    count = 0
    for i in range(len(arr)):
        count += 1
        if i == len(arr) - 1 or arr[i] != arr[i + 1]:
            result.append(count)
    
    return result

def plot_wandb_logs(df, filename, metric):
    fig, ax = plt.subplots(figsize=(8, 6)) 
    # Plot training loss for each fold
    for fold in range(1, 6):
        fold_loss = df[f'Fold {fold} - {filename}']
        ax.plot(df['epoch'], fold_loss, label=f'Fold {fold}', linewidth=2)

    ax.set_xlabel('Epoch', fontsize=14)
    ax.set_ylabel(f'{metric}', fontsize=14)
    ax.set_xlim(left=0)
    ax.set_ylim(bottom=0)
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(lambda x, pos: f"{int(x)}" if x != 0 else ""))
    ax.tick_params(axis='both', which='major', labelsize=14)
    ax.legend(fontsize=14)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    plt.savefig(f'./Manuscript/ReportImages/{filename}.png', bbox_inches='tight')
    plt.close(fig)
    
def scoring_pdiff(percentage_difference, min, max):
    scores = [(percentage / 100 * (max- min))  for percentage in percentage_difference]
    return scores

def pdiff_scoring(scores, min, max):
    percentage_differences = [(score / (max-min)) * 100 for score in scores]
    return percentage_differences

def params_analysis(dataframe, title):
    variance = dataframe.var()
    norm = Normalize(vmin=variance.min(), vmax=variance.max())
    colors = plt.cm.plasma(norm(variance.values))
    parameters = variance.index
    angles = np.linspace(0, 2 * np.pi, len(parameters), endpoint=False).tolist()
    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw={'projection': 'polar'})
    bars = ax.bar(angles, [1] * len(parameters), color=colors, alpha=0.8, width=0.3)

    for bar, angle, label in zip(bars, angles, parameters):
        rotation = np.degrees(angle)  
        label_distance = 0.6  

        ax.text(
            angle, label_distance, label, ha='center', va="center", rotation=rotation,
            rotation_mode="anchor", fontsize=8, fontweight='bold', color='red'
        )

    sm = ScalarMappable(cmap="plasma", norm=norm)
    sm.set_array([])  
    cbar = plt.colorbar(sm, ax=ax, orientation="horizontal", pad=0.1, shrink=0.8)
    cbar.set_label("Variance", fontsize=12)
    
    ax.set_yticks([])  
    ax.set_xticks([])  

    plt.savefig(f'./Manuscript/ReportImages/{title}_params.png', bbox_inches='tight')
    
def table_all_params(file_path, existing_columns):#Checks the count of paramters in the file across topology too
    try:
        df = pd.read_excel(file_path, sheet_name='input_data', header=None)
        df = df.dropna(how='all').dropna(axis=1, how='all')

        if existing_columns is None: # Tracks column names as parameters
            existing_columns = set()
        
        for _, row in df.iterrows():
            param = row[0]  
            if pd.notna(param):
                if param not in existing_columns:  
                    existing_columns.add(param)
                    
        degree_params = {param for param in existing_columns if param.startswith('deg_')}
        for degree_param in degree_params:
            radian_param = degree_param.replace('deg_', 'rad_')  # Replace 'deg_' with 'rad_'
            existing_columns.remove(degree_param)  # Remove degree-based parameter
            existing_columns.add(radian_param)  # Add radian-based parameter
            
        return existing_columns
    
    except Exception as e:
        print(f"Error processing {os.path.basename(file_path)}: {str(e)}")
        return existing_columns
    
def proportionality_params(df_selected_inputs, existing_columns):

    df_selected_inputs.rename(columns={'Unnamed: 0':'filename'}, inplace=True)
    df_selected_inputs.drop(df_selected_inputs.columns[0], axis=1, inplace=True)

    total_parameters = len(existing_columns)
    used_parameters = df_selected_inputs.shape[1] 

    labels = ['Used Parameters', 'Unused Parameters']
    sizes = [used_parameters, total_parameters - used_parameters]
    colors = ['#4CAF50', '#D3D3D3']

    plt.figure(figsize=(6, 6))
    plt.pie(
        sizes,
        # labels=labels,
        autopct='%1.1f%%',
        colors=colors,
        startangle=90,
        textprops={'fontsize': 18, 'weight': 'bold'},  # Bold text
    )
    plt.axis('equal')
    plt.gca().set_position([0, 0, 1, 1])  
    plt.savefig(f'./Manuscript/ReportImages/params_proportionality.png', bbox_inches='tight', pad_inches=0.1)